"""Registro local de personas y vinculación de múltiples casos por persona.

Los campos sensibles son opcionales y se guardan únicamente cuando la organización los carga.
"""
from __future__ import annotations
import csv, re, sqlite3, uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "asistente.db"

FIELDS = {
    "name": ("nombre", "nombre completo", "persona", "name"),
    "birth_date": ("fecha de nacimiento", "fecha_nacimiento", "nacimiento", "birth_date"),
    "age": ("edad", "age"),
    "sex_at_birth": ("sexo biologico", "sexo biológico", "sexo asignado al nacer", "sex_at_birth"),
    "gender_identity": ("identidad de genero", "identidad de género", "gender_identity"),
    "sexual_orientation": ("orientacion sexual", "orientación sexual", "sexualidad", "sexual_orientation"),
    "contact": ("contacto", "telefono", "teléfono", "celular", "email", "correo"),
    "document_id": ("dni", "documento", "document_id", "id"),
    "address": ("domicilio", "direccion", "dirección", "address"),
    "notes": ("notas", "observaciones", "notes"),
}

class PersonRegistry:
    def __init__(self, db_path: str | None = None):
        self.db_path = db_path or str(DB_PATH); Path(self.db_path).parent.mkdir(parents=True, exist_ok=True); self._init()
    def _db(self): return sqlite3.connect(self.db_path)
    def _init(self):
        with self._db() as db:
            db.execute("""CREATE TABLE IF NOT EXISTS people(
                person_id TEXT PRIMARY KEY, name TEXT NOT NULL, normalized_name TEXT NOT NULL,
                birth_date TEXT DEFAULT '', age TEXT DEFAULT '', sex_at_birth TEXT DEFAULT '',
                gender_identity TEXT DEFAULT '', sexual_orientation TEXT DEFAULT '', contact TEXT DEFAULT '',
                document_id TEXT DEFAULT '', address TEXT DEFAULT '', notes TEXT DEFAULT '',
                created_at TEXT, updated_at TEXT)""")
            db.execute("CREATE INDEX IF NOT EXISTS idx_people_name ON people(normalized_name)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_people_document ON people(document_id)")
            existing={r[1] for r in db.execute("PRAGMA table_info(cases)")}
            if "person_id" not in existing: db.execute("ALTER TABLE cases ADD COLUMN person_id TEXT DEFAULT ''")
    @staticmethod
    def normalize(value: str) -> str: return re.sub(r"\s+", " ", (value or "").strip().casefold())
    def list_people(self, query: str = ""):
        q=self.normalize(query); params=[]
        sql="SELECT person_id,name,birth_date,age,sex_at_birth,gender_identity,sexual_orientation,contact,document_id,address,notes FROM people"
        if q: sql += " WHERE normalized_name LIKE ? OR lower(document_id) LIKE ? OR lower(contact) LIKE ?"; like=f"%{q}%"; params=[like,like,like]
        sql += " ORDER BY name COLLATE NOCASE"
        with self._db() as db: rows=db.execute(sql,params).fetchall()
        keys=("person_id","name","birth_date","age","sex_at_birth","gender_identity","sexual_orientation","contact","document_id","address","notes")
        return [dict(zip(keys,r)) for r in rows]
    def get(self, person_id: str):
        rows=self.list_people(); return next((r for r in rows if r["person_id"]==person_id),None)
    def find_match(self, data: dict):
        doc=self.normalize(data.get("document_id")); name=self.normalize(data.get("name")); birth=self.normalize(data.get("birth_date"))
        with self._db() as db:
            if doc:
                row=db.execute("SELECT * FROM people WHERE lower(document_id)=? LIMIT 1",(doc,)).fetchone()
                if row:return self._row(row)
            if name and birth:
                row=db.execute("SELECT * FROM people WHERE normalized_name=? AND lower(birth_date)=? LIMIT 1",(name,birth)).fetchone()
                if row:return self._row(row)
            if name:
                rows=db.execute("SELECT * FROM people WHERE normalized_name=?",(name,)).fetchall()
                if len(rows)==1:return self._row(rows[0])
        return None
    def upsert(self, data: dict):
        clean={k:str(data.get(k,"") or "").strip() for k in FIELDS if k in data}
        if not clean.get("name"): raise ValueError("La persona necesita al menos un nombre o identificador.")
        match=self.find_match(clean); now=datetime.now().isoformat(timespec="seconds")
        if match:
            merged={**match,**{k:v for k,v in clean.items() if v}}; merged["updated_at"]=now
            with self._db() as db: db.execute("""UPDATE people SET name=?,normalized_name=?,birth_date=?,age=?,sex_at_birth=?,gender_identity=?,sexual_orientation=?,contact=?,document_id=?,address=?,notes=?,updated_at=? WHERE person_id=?""",(merged["name"],self.normalize(merged["name"]),merged["birth_date"],merged["age"],merged["sex_at_birth"],merged["gender_identity"],merged["sexual_orientation"],merged["contact"],merged["document_id"],merged["address"],merged["notes"],now,match["person_id"]))
            return match["person_id"], False
        pid=str(uuid.uuid4())[:12]
        with self._db() as db: db.execute("""INSERT INTO people VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(pid,clean["name"],self.normalize(clean["name"]),clean.get("birth_date",""),clean.get("age",""),clean.get("sex_at_birth",""),clean.get("gender_identity",""),clean.get("sexual_orientation",""),clean.get("contact",""),clean.get("document_id",""),clean.get("address",""),clean.get("notes",""),now,now))
        return pid, True
    def case_count(self, person_id: str) -> int:
        with self._db() as db:return db.execute("SELECT COUNT(*) FROM cases WHERE person_id=?",(person_id,)).fetchone()[0]
    def link_case(self, case_number: str, person_id: str):
        with self._db() as db: db.execute("UPDATE cases SET person_id=? WHERE case_number=?",(person_id,case_number))
    def cases(self, person_id: str):
        with self._db() as db: rows=db.execute("SELECT case_number,created_at,text,urgency,status,case_type FROM cases WHERE person_id=? ORDER BY created_at DESC",(person_id,)).fetchall()
        return [dict(zip(("case_number","created_at","text","urgency","status","case_type"),r)) for r in rows]
    @staticmethod
    def _row(row):
        keys=("person_id","name","normalized_name","birth_date","age","sex_at_birth","gender_identity","sexual_orientation","contact","document_id","address","notes","created_at","updated_at")
        return dict(zip(keys,row))

class PersonImporter:
    def __init__(self, registry: PersonRegistry | None = None): self.registry=registry or PersonRegistry()
    @staticmethod
    def _key(value): return re.sub(r"\s+", " ", str(value or "").strip().casefold())
    def map_row(self,row: dict):
        mapped={}
        normalized={self._key(k):v for k,v in row.items()}
        for field, aliases in FIELDS.items():
            for alias in aliases:
                if self._key(alias) in normalized:
                    mapped[field]=normalized[self._key(alias)]; break
        return mapped
    def import_xlsx(self,path):
        from openpyxl import load_workbook
        wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active; rows=list(ws.iter_rows(values_only=True));
        if not rows:return 0,0
        headers=[str(x or "") for x in rows[0]]; ok=duplicates=0
        for values in rows[1:]:
            data=self.map_row(dict(zip(headers,values)))
            if not data.get("name"):continue
            _,created=self.registry.upsert(data); ok+=1; duplicates += 0 if created else 1
        return ok,duplicates
    def import_csv(self,path):
        with open(path,"r",encoding="utf-8-sig",newline="") as f: reader=csv.DictReader(f); rows=list(reader)
        ok=duplicates=0
        for row in rows:
            data=self.map_row(row)
            if not data.get("name"):continue
            _,created=self.registry.upsert(data); ok+=1; duplicates += 0 if created else 1
        return ok,duplicates
    def import_pdf(self,path):
        from pypdf import PdfReader
        text="\n".join((p.extract_text() or "") for p in PdfReader(path).pages)
        # PDF de tablas suele variar mucho: primero intenta bloques tipo "Campo: valor".
        records=[]
        for block in re.split(r"\n\s*\n",text):
            data={}
            for line in block.splitlines():
                if ":" in line:
                    k,v=line.split(":",1); data[k]=v.strip()
            mapped=self.map_row(data)
            if mapped.get("name"):records.append(mapped)
        ok=duplicates=0
        for data in records:
            _,created=self.registry.upsert(data);ok+=1;duplicates += 0 if created else 1
        return ok,duplicates
    def import_file(self,path):
        ext=Path(path).suffix.lower()
        if ext==".xlsx":return self.import_xlsx(path)
        if ext==".csv":return self.import_csv(path)
        if ext==".pdf":return self.import_pdf(path)
        raise ValueError("Formato no compatible. Usá XLSX, CSV o PDF de texto.")
