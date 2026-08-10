"""Ingesta de documentos para biblioteca y registro de personas.

El módulo separa extracción, previsualización y persistencia. Un archivo nunca
crea/modifica personas solo por ser seleccionado: la importación de personas
requiere una acción explícita del operador.
"""
from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path
from typing import Any, Iterable

from .import_pipeline import prepare_import
from ..person_registry import PersonImporter


def _clean(text: str) -> str:
    text = re.sub(r"-\s*\n\s*", "", text or "")
    return re.sub(r"\s+", " ", text).strip()


def fingerprint(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def extract_pdf(path: str | Path) -> dict[str, Any]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = []
    for number, page in enumerate(reader.pages, 1):
        try:
            text = _clean(page.extract_text() or "")
        except Exception:
            text = ""
        pages.append({"page": number, "text": text, "characters": len(text)})
    text = _clean(" ".join(p["text"] for p in pages))
    return {
        "type": "pdf",
        "filename": Path(path).name,
        "pages": len(pages),
        "text": text,
        "characters": len(text),
        "pages_detail": pages,
        "fingerprint": fingerprint(path),
        "status": "texto_extraido" if text else "sin_texto",
    }


def preview_xlsx(path: str | Path, limit: int = 100) -> dict[str, Any]:
    from openpyxl import load_workbook

    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"type": "xlsx", "filename": Path(path).name, "headers": [], "rows": [], "row_count": 0}
    headers = [str(value or "").strip() for value in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1 : limit + 1]]
    return {
        "type": "xlsx",
        "filename": Path(path).name,
        "headers": headers,
        "rows": data,
        "row_count": max(0, len(rows) - 1),
        "preview_count": len(data),
        "fingerprint": fingerprint(path),
    }


def preview_csv(path: str | Path, limit: int = 100) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    return {
        "type": "csv",
        "filename": Path(path).name,
        "headers": reader.fieldnames or [],
        "rows": rows[:limit],
        "row_count": len(rows),
        "preview_count": min(len(rows), limit),
        "fingerprint": fingerprint(path),
    }


def preview_document(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    validation = prepare_import(str(path))
    if not validation["allowed"]:
        raise ValueError(validation["reason"])
    if path.suffix.lower() == ".pdf":
        return extract_pdf(path)
    if path.suffix.lower() == ".xlsx":
        return preview_xlsx(path)
    if path.suffix.lower() == ".csv":
        return preview_csv(path)
    raise ValueError("Formato compatible para previsualización: PDF, XLSX o CSV.")


def map_person_preview(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """Normaliza filas sin persistirlas y marca duplicados potenciales."""
    mapper = PersonImporter()
    result = []
    for row in rows:
        mapped = mapper.map_row(row)
        if mapped.get("name"):
            result.append(mapped)
    return result


def import_people_after_review(path: str | Path, registry) -> dict[str, int]:
    """Persistencia explícita posterior a la revisión del operador."""
    importer = PersonImporter(registry)
    created, duplicates = importer.import_file(str(path))
    return {"processed": created, "updated_existing": duplicates}
